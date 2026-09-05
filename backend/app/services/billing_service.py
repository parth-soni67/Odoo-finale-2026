import io
import uuid
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone, timedelta
from dateutil.relativedelta import relativedelta
from sqlalchemy.orm import Session
from fastapi import HTTPException, status

from app.models.order import Order, OrderLine, OrderStatus
from app.models.product import Product
from app.models.billing import (
    Invoice,
    InvoiceLine,
    InvoiceStatus,
    BillingType,
    Subscription,
    SubscriptionStatus,
    SubscriptionPlan,
    Payment,
    PaymentStatus,
)
from app.models.audit import AuditLog
from app.models.user import User, Role


class BillingService:
    def generate_billing(self, db: Session, order_id: int, user_id: int) -> Dict[str, Any]:
        """Generates one-time and initial recurring invoices for an order. Idempotent."""
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"code": "NOT_FOUND", "message": "Order not found"}
            )

        invoices = []
        now_utc = datetime.now(timezone.utc)

        # 1. Check for one-time order lines
        one_time_lines = [
            l for l in order.lines
            if getattr(l.line_type, "value", str(l.line_type)) == "ONE_TIME"
        ]
        if one_time_lines:
            existing_one_time = (
                db.query(Invoice)
                .filter(Invoice.order_id == order_id, Invoice.billing_type == BillingType.ONE_TIME)
                .first()
            )
            if not existing_one_time:
                subtot = round(sum(l.unit_price * l.quantity for l in one_time_lines), 2)
                disc = round(sum((l.unit_price * l.quantity) * (l.discount_percent / 100.0) for l in one_time_lines), 2)
                tot = round(subtot - disc, 2)

                invoice_number = f"INV-{uuid.uuid4().hex[:8].upper()}"
                one_time_inv = Invoice(
                    invoice_number=invoice_number,
                    order_id=order.id,
                    customer_id=order.customer_id,
                    subtotal=subtot,
                    discount=disc,
                    tax=0.0,
                    total_amount=tot,
                    currency="USD",
                    status=InvoiceStatus.DRAFT,
                    billing_type=BillingType.ONE_TIME,
                    due_date=now_utc + timedelta(days=30),
                )
                db.add(one_time_inv)
                db.flush()

                for l in one_time_lines:
                    l_sub = round(l.unit_price * l.quantity, 2)
                    l_disc = round(l_sub * (l.discount_percent / 100.0), 2)
                    l_tot = round(l_sub - l_disc, 2)
                    line_item = InvoiceLine(
                        invoice_id=one_time_inv.id,
                        product_id=l.product_id,
                        product_name=l.product.name if l.product else "One-Time Product",
                        sku=l.product.sku if l.product else None,
                        quantity=l.quantity,
                        unit_price=l.unit_price,
                        discount=l_disc,
                        line_total=l_tot,
                        billing_type=BillingType.ONE_TIME,
                    )
                    db.add(line_item)

                audit = AuditLog(
                    user_id=user_id,
                    entity_type="Invoice",
                    entity_id=one_time_inv.id,
                    action="INVOICE_CREATED",
                    new_value=f"One-time invoice created for order {order.id}",
                )
                db.add(audit)
                invoices.append(one_time_inv)

        # 2. Inspect active subscriptions tied to order
        subscriptions = db.query(Subscription).filter(Subscription.order_id == order_id).all()
        for sub in subscriptions:
            # Check expiry
            if sub.end_date:
                sub_end_cmp = sub.end_date if sub.end_date.tzinfo else sub.end_date.replace(tzinfo=timezone.utc)
                if now_utc > sub_end_cmp:
                    sub.status = SubscriptionStatus.EXPIRED

            # If active and has recurring billing frequency (not NONE)
            if sub.status == SubscriptionStatus.ACTIVE and (sub.billing_frequency or "NONE").upper() not in ("NONE", ""):
                freq = (sub.billing_frequency or "NONE").upper()
                if not sub.next_billing_date and sub.start_date and freq in ("MONTHLY", "QUARTERLY", "YEARLY"):
                    if freq == "MONTHLY":
                        sub.next_billing_date = sub.start_date + relativedelta(months=1)
                    elif freq == "QUARTERLY":
                        sub.next_billing_date = sub.start_date + relativedelta(months=3)
                    elif freq == "YEARLY":
                        sub.next_billing_date = sub.start_date + relativedelta(years=1)
                    if sub.end_date and sub.next_billing_date and sub.next_billing_date > sub.end_date:
                        sub.next_billing_date = None
                    sub.renewal_date = sub.next_billing_date

                # Check if recurring invoice already exists for this order & subscription
                existing_rec_inv = (
                    db.query(Invoice)
                    .filter(
                        Invoice.order_id == order_id,
                        Invoice.billing_type == BillingType.RECURRING,
                        Invoice.subscription_id == sub.id,
                    )
                    .first()
                )
                if not existing_rec_inv:
                    # Also check by order_id and billing_type for backwards compatibility
                    existing_any_rec = (
                        db.query(Invoice)
                        .filter(Invoice.order_id == order_id, Invoice.billing_type == BillingType.RECURRING)
                        .first()
                    )
                    if not existing_any_rec:
                        matching_line = next((l for l in order.lines if l.product_id == sub.product_id), None)
                        if matching_line:
                            line_sub = round(matching_line.unit_price * matching_line.quantity, 2)
                            line_disc = round(line_sub * (matching_line.discount_percent / 100.0), 2)
                            rec_amount = round(line_sub - line_disc, 2)
                            prod_name = matching_line.product.name if matching_line.product else (sub.name or "Subscription Item")
                            sku_val = matching_line.product.sku if matching_line.product else None
                            u_price = matching_line.unit_price
                            qty = matching_line.quantity
                        else:
                            prod = sub.product or (db.query(Product).filter(Product.id == sub.product_id).first() if sub.product_id else None)
                            rec_amount = round(prod.unit_price, 2) if prod else 0.0
                            line_sub = rec_amount
                            line_disc = 0.0
                            prod_name = prod.name if prod else (sub.name or "Subscription Item")
                            sku_val = prod.sku if prod else None
                            u_price = rec_amount
                            qty = 1

                        if rec_amount > 0:
                            rec_invoice = Invoice(
                                invoice_number=f"INV-REC-{uuid.uuid4().hex[:8].upper()}",
                                order_id=order.id,
                                customer_id=order.customer_id,
                                subscription_id=sub.id,
                                subtotal=line_sub,
                                discount=line_disc,
                                tax=0.0,
                                total_amount=rec_amount,
                                currency="USD",
                                status=InvoiceStatus.ISSUED,
                                billing_type=BillingType.RECURRING,
                                period_start=sub.start_date,
                                period_end=sub.next_billing_date or sub.end_date,
                                due_date=now_utc + timedelta(days=30),
                            )
                            db.add(rec_invoice)
                            db.flush()

                            rec_line_item = InvoiceLine(
                                invoice_id=rec_invoice.id,
                                product_id=sub.product_id,
                                subscription_id=sub.id,
                                product_name=prod_name,
                                sku=sku_val,
                                quantity=qty,
                                unit_price=u_price,
                                discount=line_disc,
                                line_total=rec_amount,
                                billing_type=BillingType.RECURRING,
                            )
                            db.add(rec_line_item)
                            invoices.append(rec_invoice)

        db.commit()

        for i in invoices:
            db.refresh(i)
        for s in subscriptions:
            db.refresh(s)

        all_order_invoices = db.query(Invoice).filter(Invoice.order_id == order_id).all()
        return {
            "invoices": all_order_invoices,
            "subscriptions": subscriptions,
        }

    def run_recurring_billing(
        self,
        db: Session,
        user_id: int,
        simulated_date: Optional[datetime] = None,
        subscription_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """Deterministic recurring billing engine.

        Finds active subscriptions, generates invoices for due periods, advances next_billing_date,
        prevents duplicates, and transitions subscriptions to EXPIRED when validity expires.
        """
        now_utc = simulated_date or datetime.now(timezone.utc)
        if now_utc.tzinfo is None:
            now_utc = now_utc.replace(tzinfo=timezone.utc)

        query = db.query(Subscription).filter(Subscription.status == SubscriptionStatus.ACTIVE)
        if subscription_id:
            query = query.filter(Subscription.id == subscription_id)
        active_subs = query.all()

        invoices_generated = []
        expired_subs = 0
        processed_count = 0

        for sub in active_subs:
            processed_count += 1
            duration_mode = (sub.duration_mode or "TILL_VALIDITY").upper()

            # Check if validity period has expired
            if duration_mode == "TILL_VALIDITY" and sub.end_date:
                end_cmp = sub.end_date if sub.end_date.tzinfo else sub.end_date.replace(tzinfo=timezone.utc)
                if simulated_date and simulated_date >= end_cmp:
                    sub.status = SubscriptionStatus.EXPIRED
                    expired_subs += 1
                    continue
                if sub.next_billing_date:
                    next_cmp = sub.next_billing_date if sub.next_billing_date.tzinfo else sub.next_billing_date.replace(tzinfo=timezone.utc)
                    if next_cmp >= end_cmp:
                        sub.status = SubscriptionStatus.EXPIRED
                        expired_subs += 1
                        continue

            freq = (sub.billing_frequency or "NONE").upper()
            if freq not in ("MONTHLY", "QUARTERLY", "YEARLY"):
                continue

            # Determine the period to invoice
            cycle_start = sub.next_billing_date or sub.current_period_end or sub.start_date or now_utc
            if cycle_start.tzinfo is None:
                cycle_start = cycle_start.replace(tzinfo=timezone.utc)

            # Check if cycle_start is at or after end_date for TILL_VALIDITY
            if duration_mode == "TILL_VALIDITY" and sub.end_date:
                end_cmp = sub.end_date if sub.end_date.tzinfo else sub.end_date.replace(tzinfo=timezone.utc)
                if cycle_start >= end_cmp:
                    sub.status = SubscriptionStatus.EXPIRED
                    expired_subs += 1
                    continue

            # Calculate cycle end
            if freq == "MONTHLY":
                cycle_end = cycle_start + relativedelta(months=1)
            elif freq == "QUARTERLY":
                cycle_end = cycle_start + relativedelta(months=3)
            elif freq == "YEARLY":
                cycle_end = cycle_start + relativedelta(years=1)
            else:
                cycle_end = cycle_start + relativedelta(months=1)

            # Idempotency check: has an invoice already been generated for this subscription and period_start?
            existing_inv = (
                db.query(Invoice)
                .filter(
                    Invoice.subscription_id == sub.id,
                    Invoice.period_start == cycle_start,
                )
                .first()
            )
            if existing_inv:
                # Already generated for this period, advance pointer if needed and skip
                continue

            # Derive product info, quantities, and pricing
            matching_line = None
            if sub.order and sub.order.lines:
                matching_line = next((l for l in sub.order.lines if l.product_id == sub.product_id), None)

            if matching_line:
                l_sub = round(float(matching_line.unit_price * matching_line.quantity), 2)
                l_disc = round(float(l_sub * (matching_line.discount_percent / 100.0)), 2)
                tot_amt = round(float(l_sub - l_disc), 2)
                prod_name = matching_line.product.name if matching_line.product else (sub.name or "Recurring Service")
                sku_val = matching_line.product.sku if matching_line.product else None
                unit_pr = matching_line.unit_price
                qty = matching_line.quantity
            else:
                prod = sub.product or (db.query(Product).filter(Product.id == sub.product_id).first() if sub.product_id else None)
                tot_amt = round(prod.unit_price, 2) if prod else 0.0
                l_sub = tot_amt
                l_disc = 0.0
                prod_name = prod.name if prod else (sub.name or "Recurring Service")
                sku_val = prod.sku if prod else None
                unit_pr = tot_amt
                qty = 1

            inv_number = f"INV-{uuid.uuid4().hex[:8].upper()}"
            invoice = Invoice(
                invoice_number=inv_number,
                order_id=sub.order_id,
                customer_id=sub.customer_id,
                subscription_id=sub.id,
                subtotal=l_sub,
                discount=l_disc,
                tax=0.0,
                total_amount=tot_amt,
                currency="USD",
                status=InvoiceStatus.ISSUED,
                billing_type=BillingType.RECURRING,
                period_start=cycle_start,
                period_end=cycle_end,
                due_date=cycle_start + timedelta(days=30),
            )
            db.add(invoice)
            db.flush()

            inv_line = InvoiceLine(
                invoice_id=invoice.id,
                product_id=sub.product_id,
                subscription_id=sub.id,
                product_name=prod_name,
                sku=sku_val,
                quantity=qty,
                unit_price=unit_pr,
                discount=l_disc,
                line_total=tot_amt,
                billing_type=BillingType.RECURRING,
            )
            db.add(inv_line)

            # Advance subscription billing cycle
            sub.current_period_start = cycle_start
            sub.current_period_end = cycle_end
            sub.next_billing_date = cycle_end
            sub.renewal_date = cycle_end

            # Check if this completed the validity period
            if duration_mode == "TILL_VALIDITY" and sub.end_date:
                end_cmp = sub.end_date if sub.end_date.tzinfo else sub.end_date.replace(tzinfo=timezone.utc)
                if sub.next_billing_date >= end_cmp:
                    sub.status = SubscriptionStatus.EXPIRED
                    sub.next_billing_date = None
                    expired_subs += 1

            audit = AuditLog(
                user_id=user_id,
                entity_type="Invoice",
                entity_id=invoice.id,
                action="RECURRING_INVOICE_GENERATED",
                new_value=f"Recurring invoice {invoice.invoice_number} generated for subscription {sub.id}",
            )
            db.add(audit)
            invoices_generated.append(invoice)

        db.commit()

        return {
            "message": f"Billing run executed. Generated {len(invoices_generated)} invoice(s).",
            "invoices_generated": len(invoices_generated),
            "subscriptions_processed": processed_count,
            "expired_subscriptions": expired_subs,
            "invoice_ids": [i.id for i in invoices_generated],
        }

    def expire_subscription(self, db: Session, subscription_id: int, user_id: int) -> Subscription:
        """Explicitly expires an active subscription and logs an audit trail."""
        sub = db.query(Subscription).filter(Subscription.id == subscription_id).first()
        if not sub:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"code": "NOT_FOUND", "message": "Subscription not found"}
            )
        sub.status = SubscriptionStatus.EXPIRED
        audit = AuditLog(
            user_id=user_id,
            entity_type="Subscription",
            entity_id=sub.id,
            action="SUBSCRIPTION_EXPIRED",
            new_value=f"Subscription '{sub.name}' expired"
        )
        db.add(audit)
        db.commit()
        db.refresh(sub)
        return sub

    def process_payment(
        self,
        db: Session,
        invoice_id: int,
        amount: float,
        payment_method: str,
        user_id: int,
    ) -> Payment:
        invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not invoice:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"code": "NOT_FOUND", "message": "Invoice not found"},
            )

        if invoice.status == InvoiceStatus.PAID:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"code": "ALREADY_PAID", "message": "Invoice is already paid"},
            )

        if round(amount, 2) != round(invoice.total_amount, 2):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "code": "INVALID_AMOUNT",
                    "message": f"Payment amount ({amount}) must match invoice total ({invoice.total_amount})",
                },
            )

        payment = Payment(
            invoice_id=invoice.id,
            amount=amount,
            payment_method=payment_method or "SIMULATED_CARD",
            payment_status=PaymentStatus.SUCCESSFUL,
            transaction_id=f"TXN-{uuid.uuid4().hex[:12].upper()}",
        )
        db.add(payment)

        invoice.status = InvoiceStatus.PAID

        audit_pay = AuditLog(
            user_id=user_id,
            entity_type="Payment",
            entity_id=invoice.id,
            action="PAYMENT_COMPLETED",
            new_value=f"Payment of {amount} completed for invoice {invoice.id}",
        )
        db.add(audit_pay)

        db.commit()
        db.refresh(payment)
        return payment

    def get_billing_summary(self, db: Session, order_id: int) -> Dict[str, Any]:
        invoices = db.query(Invoice).filter(Invoice.order_id == order_id).all()
        subscriptions = db.query(Subscription).filter(Subscription.order_id == order_id).all()
        return {
            "invoices": invoices,
            "subscriptions": subscriptions,
        }

    def generate_invoice_pdf(self, invoice: Invoice) -> bytes:
        """Generates a professional enterprise PDF document for the invoice using ReportLab."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#1E3A8A'),
            fontName='Helvetica-Bold',
        )
        section_heading = ParagraphStyle(
            'SecHead',
            parent=styles['Heading2'],
            fontSize=11,
            leading=15,
            textColor=colors.HexColor('#1F2937'),
            fontName='Helvetica-Bold',
        )
        body_style = ParagraphStyle(
            'Body',
            parent=styles['Normal'],
            fontSize=9,
            leading=13,
            textColor=colors.HexColor('#374151'),
        )
        bold_body = ParagraphStyle(
            'BoldBody',
            parent=styles['Normal'],
            fontSize=9,
            leading=13,
            textColor=colors.HexColor('#111827'),
            fontName='Helvetica-Bold',
        )

        story = []

        # Top Header
        header_data = [
            [
                Paragraph("<b>DEALFLOW360</b><br/><font size=8 color='#6B7280'>Intelligent Sales Operations & Billing Platform</font>", title_style),
                Paragraph(f"<b>INVOICE</b><br/><font size=11 color='#2563EB'><b>{invoice.invoice_number}</b></font>", ParagraphStyle('RHead', parent=title_style, alignment=2)),
            ]
        ]
        t_header = Table(header_data, colWidths=[340, 200])
        t_header.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(t_header)
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#2563EB'), spaceAfter=14))

        # Metadata & Bill-To Info Box
        cust = invoice.customer
        company_name = cust.company_name if cust else "Valued Customer"
        contact_name = cust.contact_name if cust else "Finance Dept"
        cust_email = cust.email if cust else "billing@customer.com"
        order_num = invoice.order.order_number if invoice.order else (f"ORD-{invoice.order_id}" if invoice.order_id else "N/A")

        bill_to_html = f"<b>BILL TO:</b><br/><b>{company_name}</b><br/>Attn: {contact_name}<br/>Email: {cust_email}"
        inv_date_str = invoice.created_at.strftime('%d %b %Y') if invoice.created_at else "N/A"
        due_date_str = invoice.due_date.strftime('%d %b %Y') if invoice.due_date else "Due on Receipt"
        status_str = invoice.status.value if hasattr(invoice.status, 'value') else str(invoice.status)
        type_str = invoice.billing_type.value if hasattr(invoice.billing_type, 'value') else str(invoice.billing_type)

        meta_html = (
            f"<b>Invoice Date:</b> {inv_date_str}<br/>"
            f"<b>Due Date:</b> {due_date_str}<br/>"
            f"<b>Order Number:</b> {order_num}<br/>"
            f"<b>Invoice Type:</b> {type_str}<br/>"
            f"<b>Status:</b> <b>{status_str}</b>"
        )

        info_table = Table(
            [[Paragraph(bill_to_html, body_style), Paragraph(meta_html, body_style)]],
            colWidths=[320, 220],
        )
        info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 12))

        # Subscription banner if applicable
        sub = invoice.subscription
        if sub or type_str == "RECURRING":
            sub_name = sub.name if sub else "Recurring Subscription Plan"
            freq = sub.billing_frequency if sub else "MONTHLY"
            d_mode = sub.duration_mode if sub else "TILL_VALIDITY"
            p_start = invoice.period_start.strftime('%d %b %Y') if invoice.period_start else "N/A"
            p_end = invoice.period_end.strftime('%d %b %Y') if invoice.period_end else "N/A"
            sub_info_html = (
                f"<b>Subscription:</b> {sub_name} &nbsp;|&nbsp; "
                f"<b>Cadence:</b> {freq} &nbsp;|&nbsp; "
                f"<b>Duration:</b> {d_mode} &nbsp;|&nbsp; "
                f"<b>Period:</b> {p_start} &rarr; {p_end}"
            )
            sub_table = Table([[Paragraph(sub_info_html, body_style)]], colWidths=[540])
            sub_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EFF6FF')),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#93C5FD')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(sub_table)
            story.append(Spacer(1, 12))

        # Line Items Table
        lines = invoice.lines
        lines_data = [
            [
                Paragraph("<b>Product</b>", bold_body),
                Paragraph("<b>SKU</b>", bold_body),
                Paragraph("<b>Qty</b>", bold_body),
                Paragraph("<b>Unit Price</b>", bold_body),
                Paragraph("<b>Discount</b>", bold_body),
                Paragraph("<b>Line Total</b>", bold_body),
            ]
        ]

        if lines:
            for l in lines:
                lines_data.append([
                    Paragraph(l.product_name or "Item", body_style),
                    Paragraph(l.sku or "-", body_style),
                    Paragraph(str(l.quantity), body_style),
                    Paragraph(f"${l.unit_price:,.2f}", body_style),
                    Paragraph(f"${l.discount:,.2f}", body_style),
                    Paragraph(f"${l.line_total:,.2f}", bold_body),
                ])
        elif invoice.order and invoice.order.lines:
            for l in invoice.order.lines:
                l_disc = round(float(l.unit_price * l.quantity * (l.discount_percent / 100.0)), 2)
                lines_data.append([
                    Paragraph(l.product.name if l.product else "Item", body_style),
                    Paragraph(l.product.sku if l.product else "-", body_style),
                    Paragraph(str(l.quantity), body_style),
                    Paragraph(f"${l.unit_price:,.2f}", body_style),
                    Paragraph(f"${l_disc:,.2f}", body_style),
                    Paragraph(f"${l.line_total:,.2f}", bold_body),
                ])
        else:
            lines_data.append([
                Paragraph(f"Invoice {invoice.invoice_number}", body_style),
                Paragraph("-", body_style),
                Paragraph("1", body_style),
                Paragraph(f"${invoice.total_amount:,.2f}", body_style),
                Paragraph("$0.00", body_style),
                Paragraph(f"${invoice.total_amount:,.2f}", bold_body),
            ])

        items_table = Table(lines_data, colWidths=[200, 70, 40, 75, 75, 80])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 12))

        # Totals Summary
        subtot = invoice.subtotal if invoice.subtotal > 0 else invoice.total_amount
        disc = invoice.discount
        tax = invoice.tax
        grand_tot = invoice.total_amount

        totals_data = [
            [Paragraph("Subtotal:", body_style), Paragraph(f"${subtot:,.2f}", body_style)],
            [Paragraph("Discount:", body_style), Paragraph(f"-${disc:,.2f}", body_style)],
            [Paragraph("Tax:", body_style), Paragraph(f"${tax:,.2f}", body_style)],
            [Paragraph("<b>Total Amount:</b>", bold_body), Paragraph(f"<b>${grand_tot:,.2f} {invoice.currency or 'USD'}</b>", bold_body)],
        ]
        totals_table = Table(totals_data, colWidths=[120, 100])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LINEBELOW', (0, 2), (1, 2), 0.5, colors.HexColor('#9CA3AF')),
            ('BACKGROUND', (0, 3), (1, 3), colors.HexColor('#F3F4F6')),
        ]))
        wrapper_table = Table([["", totals_table]], colWidths=[320, 220])
        wrapper_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(wrapper_table)
        story.append(Spacer(1, 16))

        # Payment Status Box
        payments = invoice.payments
        if payments and invoice.status == InvoiceStatus.PAID:
            latest_pay = payments[-1]
            pay_info = (
                f"<font color='#059669'><b>PAID IN FULL</b></font> &nbsp;|&nbsp; "
                f"Transaction ID: {latest_pay.transaction_id or 'TXN-SUCCESS'} &nbsp;|&nbsp; "
                f"Method: {latest_pay.payment_method} &nbsp;|&nbsp; "
                f"Paid: {latest_pay.created_at.strftime('%d %b %Y') if latest_pay.created_at else ''}"
            )
        else:
            pay_info = f"<b>Payment Status:</b> <font color='#D97706'><b>{status_str}</b></font> &nbsp;|&nbsp; Payment Due: {due_date_str}"

        pay_table = Table([[Paragraph(pay_info, body_style)]], colWidths=[540])
        pay_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(pay_table)
        story.append(Spacer(1, 20))

        # Footer
        footer_text = "<font size=8 color='#9CA3AF'>DealFlow360 Enterprise Billing &middot; Questions? Contact support@dealflow360.com &middot; Retain this invoice for your records.</font>"
        story.append(Paragraph(footer_text, ParagraphStyle('Foot', parent=body_style, alignment=1)))

        doc.build(story)
        return buffer.getvalue()

    def generate_invoice_xlsx(self, invoice: Invoice) -> bytes:
        """Generates an Excel spreadsheet (.xlsx) for the invoice with structured data using openpyxl."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        title_font = Font(name="Calibri", size=15, bold=True, color="1E3A8A")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        section_font = Font(name="Calibri", size=11, bold=True, color="1F2937")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB'),
        )

        ws["A1"] = "DealFlow360 — Invoice"
        ws["A1"].font = title_font
        ws["A2"] = f"Official Invoice: {invoice.invoice_number}"
        ws["A2"].font = bold_font

        ws["A4"] = "Invoice Information"
        ws["A4"].font = section_font

        inv_type_val = invoice.billing_type.value if hasattr(invoice.billing_type, 'value') else str(invoice.billing_type)
        status_val = invoice.status.value if hasattr(invoice.status, 'value') else str(invoice.status)

        info_rows = [
            ("Invoice Number", invoice.invoice_number),
            ("Invoice Date", invoice.created_at.strftime("%Y-%m-%d") if invoice.created_at else "N/A"),
            ("Due Date", invoice.due_date.strftime("%Y-%m-%d") if invoice.due_date else "N/A"),
            ("Invoice Type", inv_type_val),
            ("Status", status_val),
            ("Currency", invoice.currency or "USD"),
            ("Order Number", invoice.order.order_number if invoice.order else (f"ORD-{invoice.order_id}" if invoice.order_id else "N/A")),
            ("Customer Company", invoice.customer.company_name if invoice.customer else "N/A"),
            ("Customer Contact", invoice.customer.contact_name if invoice.customer else "N/A"),
            ("Customer Email", invoice.customer.email if invoice.customer else "N/A"),
        ]

        curr_row = 5
        for label, val in info_rows:
            ws.cell(row=curr_row, column=1, value=label).font = bold_font
            ws.cell(row=curr_row, column=2, value=val).font = regular_font
            curr_row += 1

        if invoice.subscription or inv_type_val == "RECURRING":
            curr_row += 1
            ws.cell(row=curr_row, column=1, value="Subscription Information").font = section_font
            curr_row += 1
            sub = invoice.subscription
            sub_rows = [
                ("Subscription Plan", sub.name if sub else "Recurring Subscription"),
                ("Billing Frequency", sub.billing_frequency if sub else "MONTHLY"),
                ("Duration Mode", sub.duration_mode if sub else "TILL_VALIDITY"),
                ("Period Start", invoice.period_start.strftime("%Y-%m-%d") if invoice.period_start else "N/A"),
                ("Period End", invoice.period_end.strftime("%Y-%m-%d") if invoice.period_end else "N/A"),
            ]
            for label, val in sub_rows:
                ws.cell(row=curr_row, column=1, value=label).font = bold_font
                ws.cell(row=curr_row, column=2, value=val).font = regular_font
                curr_row += 1

        curr_row += 1
        ws.cell(row=curr_row, column=1, value="Invoice Lines").font = section_font
        curr_row += 1

        headers = ["Product", "SKU", "Quantity", "Unit Price", "Discount", "Line Total", "Type"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if col_idx in [3, 7] else ("right" if col_idx in [4, 5, 6] else "left"))
        curr_row += 1

        lines = invoice.lines
        if lines:
            for l in lines:
                l_type_str = l.billing_type.value if hasattr(l.billing_type, 'value') else str(l.billing_type)
                row_vals = [
                    l.product_name or "Item",
                    l.sku or "-",
                    l.quantity,
                    l.unit_price,
                    l.discount,
                    l.line_total,
                    l_type_str,
                ]
                for col_idx, v in enumerate(row_vals, start=1):
                    c = ws.cell(row=curr_row, column=col_idx, value=v)
                    c.font = regular_font
                    c.border = thin_border
                    if col_idx in [4, 5, 6]:
                        c.number_format = "$#,##0.00"
                curr_row += 1
        elif invoice.order and invoice.order.lines:
            for l in invoice.order.lines:
                l_disc = round(float(l.unit_price * l.quantity * (l.discount_percent / 100.0)), 2)
                row_vals = [
                    l.product.name if l.product else "Item",
                    l.product.sku if l.product else "-",
                    l.quantity,
                    l.unit_price,
                    l_disc,
                    l.line_total,
                    getattr(l.line_type, 'value', str(l.line_type)),
                ]
                for col_idx, v in enumerate(row_vals, start=1):
                    c = ws.cell(row=curr_row, column=col_idx, value=v)
                    c.font = regular_font
                    c.border = thin_border
                    if col_idx in [4, 5, 6]:
                        c.number_format = "$#,##0.00"
                curr_row += 1
        else:
            row_vals = [f"Invoice {invoice.invoice_number}", "-", 1, invoice.total_amount, 0.0, invoice.total_amount, "ONE_TIME"]
            for col_idx, v in enumerate(row_vals, start=1):
                c = ws.cell(row=curr_row, column=col_idx, value=v)
                c.font = regular_font
                c.border = thin_border
                if col_idx in [4, 5, 6]:
                    c.number_format = "$#,##0.00"
            curr_row += 1

        curr_row += 1
        subtot = invoice.subtotal if invoice.subtotal > 0 else invoice.total_amount
        disc = invoice.discount
        tax = invoice.tax
        grand_tot = invoice.total_amount

        totals_rows = [
            ("Subtotal", subtot),
            ("Discount", disc),
            ("Tax", tax),
            ("Grand Total", grand_tot),
        ]
        for label, val in totals_rows:
            ws.cell(row=curr_row, column=5, value=label).font = bold_font
            c = ws.cell(row=curr_row, column=6, value=val)
            c.font = bold_font if label == "Grand Total" else regular_font
            c.number_format = "$#,##0.00"
            curr_row += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


billing_service = BillingService()
