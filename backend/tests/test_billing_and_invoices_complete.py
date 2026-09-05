import io
import pytest
from datetime import datetime, timezone, timedelta
from dateutil.relativedelta import relativedelta
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
import openpyxl

from app.models.user import User, Role
from app.models.customer import Customer, CustomerTier
from app.models.product import Product, ProductCategory, FulfillmentType
from app.models.quote import Quote, QuoteLine, QuoteStatus, LineType
from app.models.order import Order, OrderLine, OrderStatus, FulfillmentSplit
from app.models.billing import Invoice, InvoiceLine, InvoiceStatus, BillingType, Subscription, SubscriptionStatus, Payment, PaymentStatus
from app.models.warehouse import Warehouse, Inventory


@pytest.fixture
def auth_tokens(client: TestClient):
    """Provides tokens for various roles."""
    resp_admin = client.post("/api/auth/login", json={"email": "admin@dealflow360.internal", "password": "Demo1234!"})
    resp_finance = client.post("/api/auth/login", json={"email": "finance@dealflow360.internal", "password": "Demo1234!"})
    resp_salesrep = client.post("/api/auth/login", json={"email": "salesrep@dealflow360.internal", "password": "Demo1234!"})
    resp_cust_a = client.post("/api/auth/login", json={"email": "customer@acmecorp.com", "password": "Demo1234!"})

    return {
        "admin": resp_admin.json()["access_token"],
        "finance": resp_finance.json()["access_token"],
        "salesrep": resp_salesrep.json()["access_token"],
        "customer_a": resp_cust_a.json()["access_token"],
    }


@pytest.fixture
def customer_b_setup(db_session: Session, client: TestClient):
    """Creates a second customer (Customer B) and returns token."""
    from app.core.security import get_password_hash
    user_b = db_session.query(User).filter(User.email == "user@technova.com").first()
    if not user_b:
        user_b = User(
            email="user@technova.com",
            hashed_password=get_password_hash("Demo1234!"),
            full_name="TechNova User",
            role=Role.CUSTOMER,
            is_active=True,
        )
        db_session.add(user_b)
        db_session.flush()

    cust_b = db_session.query(Customer).filter(Customer.email == "contact@technova.com").first()
    if not cust_b:
        cust_b = Customer(
            company_name="TechNova Corp",
            contact_name="TechNova User",
            email="contact@technova.com",
            tier=CustomerTier.GROWTH,
        )
        db_session.add(cust_b)
        db_session.flush()

    # Link user_b email domain to cust_b
    user_b.email = "procurement@technova.com"
    cust_b.email = "procurement@technova.com"
    db_session.commit()

    resp = client.post("/api/auth/login", json={"email": "procurement@technova.com", "password": "Demo1234!"})
    return {
        "token": resp.json()["access_token"],
        "customer": cust_b,
        "user": user_b,
    }


# =========================================================================
# 1. ONE-TIME INVOICE CREATION
# =========================================================================
def test_1_one_time_invoice_creation(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    laptop = db_session.query(Product).filter(Product.sku == "HW-IOT-100").first()

    order = Order(
        order_number="ORD-TEST-OT-001",
        customer_id=customer.id,
        status=OrderStatus.CONFIRMED,
        total_amount=2160.0,
    )
    db_session.add(order)
    db_session.flush()

    order_line = OrderLine(
        order_id=order.id,
        product_id=laptop.id,
        quantity=2,
        unit_price=1200.0,
        discount_percent=10.0,
        line_total=2160.0,
        line_type=LineType.ONE_TIME,
    )
    db_session.add(order_line)
    db_session.commit()

    res = client.post(
        f"/api/orders/{order.id}/billing",
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert res.status_code == 200
    data = res.json()
    assert len(data["invoices"]) >= 1

    inv = next(i for i in data["invoices"] if i["billing_type"] == "ONE_TIME")
    assert inv["total_amount"] == 2160.0
    assert inv["order_id"] == order.id

    # Verify lines in detail endpoint
    det = client.get(f"/api/invoices/{inv['id']}", headers={"Authorization": f"Bearer {auth_tokens['finance']}"})
    assert det.status_code == 200
    det_data = det.json()
    assert len(det_data["lines"]) >= 1
    assert det_data["lines"][0]["quantity"] == 2
    assert det_data["lines"][0]["unit_price"] == 1200.0


# =========================================================================
# 2. RECURRING SUBSCRIPTION INVOICE CREATION
# =========================================================================
def test_2_recurring_subscription_invoice_creation(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SUB-SUPP-247").first()

    order = Order(
        order_number="ORD-TEST-REC-001",
        customer_id=customer.id,
        status=OrderStatus.CONFIRMED,
        total_amount=500.0,
    )
    db_session.add(order)
    db_session.flush()

    order_line = OrderLine(
        order_id=order.id,
        product_id=software.id,
        quantity=1,
        unit_price=500.0,
        discount_percent=0.0,
        line_total=500.0,
        line_type=LineType.RECURRING,
        subscription_enabled=True,
        subscription_name="Cloud Support SLA",
        duration_mode="TILL_VALIDITY",
        validity_value=3,
        validity_unit="MONTHS",
        billing_frequency="MONTHLY",
    )
    db_session.add(order_line)
    db_session.flush()

    sub = Subscription(
        customer_id=customer.id,
        order_id=order.id,
        product_id=software.id,
        name="Cloud Support SLA",
        duration_mode="TILL_VALIDITY",
        validity_value=3,
        validity_unit="MONTHS",
        billing_frequency="MONTHLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=datetime.now(timezone.utc),
        end_date=datetime.now(timezone.utc) + relativedelta(months=3),
        next_billing_date=datetime.now(timezone.utc) + relativedelta(months=1),
    )
    db_session.add(sub)
    db_session.commit()

    res = client.post(
        f"/api/orders/{order.id}/billing",
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert res.status_code == 200
    data = res.json()
    rec_invs = [i for i in data["invoices"] if i["billing_type"] == "RECURRING"]
    assert len(rec_invs) == 1
    assert rec_invs[0]["total_amount"] == 500.0
    assert rec_invs[0]["subscription_id"] == sub.id


# =========================================================================
# 3. MONTHLY BILLING
# =========================================================================
def test_3_monthly_billing(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    start_dt = datetime(2026, 9, 1, 0, 0, 0, tzinfo=timezone.utc)
    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="Monthly SaaS",
        duration_mode="LIFETIME",
        billing_frequency="MONTHLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=start_dt,
        next_billing_date=datetime(2026, 10, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add(sub)
    db_session.commit()

    run_res = client.post(
        "/api/billing/run",
        json={"subscription_id": sub.id},
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert run_res.status_code == 200
    assert run_res.json()["invoices_generated"] == 1

    db_session.refresh(sub)
    # Next billing date should advance to November 1
    assert sub.next_billing_date.month == 11
    assert sub.status == SubscriptionStatus.ACTIVE


# =========================================================================
# 4. QUARTERLY BILLING
# =========================================================================
def test_4_quarterly_billing(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    start_dt = datetime(2026, 9, 1, 0, 0, 0, tzinfo=timezone.utc)
    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="Quarterly Enterprise License",
        duration_mode="LIFETIME",
        billing_frequency="QUARTERLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=start_dt,
        next_billing_date=datetime(2026, 12, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add(sub)
    db_session.commit()

    run_res = client.post(
        "/api/billing/run",
        json={"subscription_id": sub.id},
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert run_res.status_code == 200
    assert run_res.json()["invoices_generated"] == 1

    db_session.refresh(sub)
    # 2026-12-01 + 3 months -> 2027-03-01
    assert sub.next_billing_date.month == 3
    assert sub.next_billing_date.year == 2027


# =========================================================================
# 5. YEARLY BILLING
# =========================================================================
def test_5_yearly_billing(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    start_dt = datetime(2026, 9, 1, 0, 0, 0, tzinfo=timezone.utc)
    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="Annual Subscription",
        duration_mode="LIFETIME",
        billing_frequency="YEARLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=start_dt,
        next_billing_date=datetime(2027, 9, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add(sub)
    db_session.commit()

    run_res = client.post(
        "/api/billing/run",
        json={"subscription_id": sub.id},
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert run_res.status_code == 200
    assert run_res.json()["invoices_generated"] == 1

    db_session.refresh(sub)
    # Advances 1 year to 2028
    assert sub.next_billing_date.year == 2028


# =========================================================================
# 6. LIFETIME SUBSCRIPTION
# =========================================================================
def test_6_lifetime_subscription(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="Lifetime Tier Subscription",
        duration_mode="LIFETIME",
        billing_frequency="MONTHLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=datetime(2026, 9, 1, 0, 0, 0, tzinfo=timezone.utc),
        end_date=None,
        next_billing_date=datetime(2026, 10, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add(sub)
    db_session.commit()

    # Lifetime has no expiration date, multiple billing runs keep advancing
    for _ in range(3):
        res = client.post(
            "/api/billing/run",
            json={"subscription_id": sub.id},
            headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
        )
        assert res.status_code == 200

    db_session.refresh(sub)
    assert sub.status == SubscriptionStatus.ACTIVE
    assert sub.end_date is None


# =========================================================================
# 7. TILL_VALIDITY SUBSCRIPTION
# =========================================================================
def test_7_till_validity_subscription(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    start_dt = datetime(2026, 9, 1, 0, 0, 0, tzinfo=timezone.utc)
    end_dt = datetime(2026, 12, 1, 0, 0, 0, tzinfo=timezone.utc)

    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="3-Month Validity Software",
        duration_mode="TILL_VALIDITY",
        validity_value=3,
        validity_unit="MONTHS",
        billing_frequency="MONTHLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=start_dt,
        end_date=end_dt,
        next_billing_date=datetime(2026, 10, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add(sub)
    db_session.commit()

    # Run 1: invoices for Oct
    res1 = client.post("/api/billing/run", json={"subscription_id": sub.id}, headers={"Authorization": f"Bearer {auth_tokens['finance']}"})
    assert res1.json()["invoices_generated"] == 1
    db_session.refresh(sub)
    assert sub.status == SubscriptionStatus.ACTIVE

    # Run 2: invoices for Nov
    res2 = client.post("/api/billing/run", json={"subscription_id": sub.id}, headers={"Authorization": f"Bearer {auth_tokens['finance']}"})
    assert res2.json()["invoices_generated"] == 1
    db_session.refresh(sub)
    # Next billing reached or passed 2026-12-01 end date -> transitions to EXPIRED
    assert sub.status == SubscriptionStatus.EXPIRED


# =========================================================================
# 8. SUBSCRIPTION EXPIRY
# =========================================================================
def test_8_subscription_expiry(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="Expiring License",
        duration_mode="TILL_VALIDITY",
        billing_frequency="MONTHLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=datetime(2026, 5, 1, 0, 0, 0, tzinfo=timezone.utc),
        end_date=datetime(2026, 8, 1, 0, 0, 0, tzinfo=timezone.utc),
        next_billing_date=datetime(2026, 8, 1, 0, 0, 0, tzinfo=timezone.utc),
    )
    db_session.add(sub)
    db_session.commit()

    res = client.post(
        "/api/billing/run",
        json={"subscription_id": sub.id},
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert res.status_code == 200
    db_session.refresh(sub)
    assert sub.status == SubscriptionStatus.EXPIRED


# =========================================================================
# 9. NO INVOICE AFTER EXPIRY
# =========================================================================
def test_9_no_invoice_after_expiry(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    sub = Subscription(
        customer_id=customer.id,
        product_id=software.id,
        name="Already Expired Subscription",
        duration_mode="TILL_VALIDITY",
        status=SubscriptionStatus.EXPIRED,
        start_date=datetime(2026, 1, 1, 0, 0, 0, tzinfo=timezone.utc),
        end_date=datetime(2026, 4, 1, 0, 0, 0, tzinfo=timezone.utc),
        next_billing_date=None,
    )
    db_session.add(sub)
    db_session.commit()

    res = client.post(
        "/api/billing/run",
        json={"subscription_id": sub.id},
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert res.status_code == 200
    assert res.json()["invoices_generated"] == 0


# =========================================================================
# 10. DUPLICATE BILLING PREVENTION (IDEMPOTENCY)
# =========================================================================
def test_10_duplicate_billing_prevention(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    laptop = db_session.query(Product).filter(Product.sku == "HW-IOT-100").first()

    order = Order(
        order_number="ORD-TEST-IDEMPOTENT-001",
        customer_id=customer.id,
        status=OrderStatus.CONFIRMED,
        total_amount=1200.0,
    )
    db_session.add(order)
    db_session.flush()

    line = OrderLine(
        order_id=order.id,
        product_id=laptop.id,
        quantity=1,
        unit_price=1200.0,
        discount_percent=0.0,
        line_total=1200.0,
        line_type=LineType.ONE_TIME,
    )
    db_session.add(line)
    db_session.commit()

    # Call 1
    res1 = client.post(f"/api/orders/{order.id}/billing", headers={"Authorization": f"Bearer {auth_tokens['finance']}"})
    assert res1.status_code == 200
    count_1 = len(res1.json()["invoices"])

    # Call 2
    res2 = client.post(f"/api/orders/{order.id}/billing", headers={"Authorization": f"Bearer {auth_tokens['finance']}"})
    assert res2.status_code == 200
    count_2 = len(res2.json()["invoices"])

    # Invoices count must not double
    assert count_1 == count_2


# =========================================================================
# 11. PAYMENT -> INVOICE PAID
# =========================================================================
def test_11_payment_marks_invoice_paid(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()

    invoice = Invoice(
        invoice_number="INV-PAY-TEST-001",
        customer_id=customer.id,
        subtotal=800.0,
        discount=0.0,
        total_amount=800.0,
        status=InvoiceStatus.ISSUED,
        billing_type=BillingType.ONE_TIME,
    )
    db_session.add(invoice)
    db_session.commit()

    pay_res = client.post(
        f"/api/invoices/{invoice.id}/payment",
        json={"invoice_id": invoice.id, "amount": 800.0, "payment_method": "SIMULATED_CARD"},
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert pay_res.status_code == 200
    assert pay_res.json()["payment_status"] == "SUCCESSFUL"

    db_session.refresh(invoice)
    assert invoice.status == InvoiceStatus.PAID
    assert len(invoice.payments) == 1


# =========================================================================
# 12. CUSTOMER INVOICE ACCESS
# =========================================================================
def test_12_customer_invoice_access(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()

    invoice = Invoice(
        invoice_number="INV-CUST-OWN-001",
        customer_id=customer.id,
        subtotal=150.0,
        total_amount=150.0,
        status=InvoiceStatus.ISSUED,
        billing_type=BillingType.ONE_TIME,
    )
    db_session.add(invoice)
    db_session.commit()

    # Customer A should be able to view their own invoice
    res = client.get(
        f"/api/invoices/{invoice.id}",
        headers={"Authorization": f"Bearer {auth_tokens['customer_a']}"},
    )
    assert res.status_code == 200
    assert res.json()["invoice_number"] == "INV-CUST-OWN-001"


# =========================================================================
# 13. CROSS-CUSTOMER INVOICE RETURNS 403
# =========================================================================
def test_13_cross_customer_invoice_returns_403(client: TestClient, db_session: Session, auth_tokens, customer_b_setup):
    customer_a = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()

    invoice_a = Invoice(
        invoice_number="INV-CUST-A-SECRET",
        customer_id=customer_a.id,
        subtotal=999.0,
        total_amount=999.0,
        status=InvoiceStatus.ISSUED,
        billing_type=BillingType.ONE_TIME,
    )
    db_session.add(invoice_a)
    db_session.commit()

    # Customer B attempts to access Customer A's invoice -> must return 403 Forbidden!
    res = client.get(
        f"/api/invoices/{invoice_a.id}",
        headers={"Authorization": f"Bearer {customer_b_setup['token']}"},
    )
    assert res.status_code == 403

    # Also check PDF export returns 403
    res_pdf = client.get(
        f"/api/invoices/{invoice_a.id}/pdf",
        headers={"Authorization": f"Bearer {customer_b_setup['token']}"},
    )
    assert res_pdf.status_code == 403

    # Also check XLSX export returns 403
    res_xlsx = client.get(
        f"/api/invoices/{invoice_a.id}/xlsx",
        headers={"Authorization": f"Bearer {customer_b_setup['token']}"},
    )
    assert res_xlsx.status_code == 403


# =========================================================================
# 14. PDF GENERATION
# =========================================================================
def test_14_pdf_generation(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()

    inv = Invoice(
        invoice_number="INV-PDF-VERIFY-001",
        customer_id=customer.id,
        subtotal=1200.0,
        discount=120.0,
        tax=0.0,
        total_amount=1080.0,
        status=InvoiceStatus.ISSUED,
        billing_type=BillingType.ONE_TIME,
    )
    db_session.add(inv)
    db_session.flush()

    line = InvoiceLine(
        invoice_id=inv.id,
        product_name="Enterprise Laptop",
        sku="HW-LAPTOP-X1",
        quantity=1,
        unit_price=1200.0,
        discount=120.0,
        line_total=1080.0,
    )
    db_session.add(line)
    db_session.commit()

    res = client.get(
        f"/api/invoices/{inv.id}/pdf",
        headers={"Authorization": f"Bearer {auth_tokens['customer_a']}"},
    )
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    assert res.content.startswith(b"%PDF")
    assert len(res.content) > 500


# =========================================================================
# 15. XLSX GENERATION
# =========================================================================
def test_15_xlsx_generation(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()

    inv = Invoice(
        invoice_number="INV-XLSX-VERIFY-001",
        customer_id=customer.id,
        subtotal=800.0,
        discount=0.0,
        total_amount=800.0,
        status=InvoiceStatus.ISSUED,
        billing_type=BillingType.RECURRING,
    )
    db_session.add(inv)
    db_session.commit()

    res = client.get(
        f"/api/invoices/{inv.id}/xlsx",
        headers={"Authorization": f"Bearer {auth_tokens['customer_a']}"},
    )
    assert res.status_code == 200
    assert "openxmlformats" in res.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws["A1"].value == "DealFlow360 — Invoice"
    assert "INV-XLSX-VERIFY-001" in str(ws["B5"].value or ws["A2"].value)


# =========================================================================
# 16. INVOICE HISTORY PERSISTENCE (IMMUTABILITY)
# =========================================================================
def test_16_invoice_history_persistence(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    product = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()

    inv = Invoice(
        invoice_number="INV-HIST-IMMUTABLE",
        customer_id=customer.id,
        subtotal=800.0,
        total_amount=800.0,
        status=InvoiceStatus.PAID,
        billing_type=BillingType.RECURRING,
    )
    db_session.add(inv)
    db_session.commit()

    # Even if product price increases later
    orig_price = product.unit_price
    product.unit_price = orig_price + 200.0
    db_session.commit()

    # Historical invoice must retain original amount $800.0
    db_session.refresh(inv)
    assert inv.total_amount == 800.0

    # Restore price
    product.unit_price = orig_price
    db_session.commit()


# =========================================================================
# 17. PHYSICAL PRODUCT FULFILLMENT
# =========================================================================
def test_17_physical_product_fulfillment(client: TestClient, db_session: Session, auth_tokens):
    laptop = db_session.query(Product).filter(Product.sku == "HW-IOT-100").first()
    assert getattr(laptop, "fulfillment_type", "PHYSICAL") == "PHYSICAL"

    customer = db_session.query(Customer).first()
    order = Order(
        order_number="ORD-TEST-PHYSICAL-001",
        customer_id=customer.id,
        status=OrderStatus.PENDING,
        total_amount=laptop.unit_price,
    )
    db_session.add(order)
    db_session.flush()

    line = OrderLine(
        order_id=order.id,
        product_id=laptop.id,
        quantity=1,
        unit_price=laptop.unit_price,
        line_total=laptop.unit_price,
        line_type=LineType.ONE_TIME,
    )
    db_session.add(line)
    db_session.commit()

    from app.services.fulfillment_service import fulfillment_service
    sugg = fulfillment_service.suggest_fulfillment(db_session, order.id)
    # Physical product should have suggestions for warehouse inventory
    assert len(sugg["lines"]) == 1
    assert sugg["lines"][0]["product_id"] == laptop.id


# =========================================================================
# 18. DIGITAL PRODUCT WITHOUT WAREHOUSE REQUIREMENT
# =========================================================================
def test_18_digital_product_without_warehouse_requirement(client: TestClient, db_session: Session, auth_tokens):
    software = db_session.query(Product).filter(Product.sku == "SW-DF360-LIC").first()
    software.fulfillment_type = "DIGITAL"
    db_session.commit()

    customer = db_session.query(Customer).first()
    order = Order(
        order_number="ORD-TEST-DIGITAL-001",
        customer_id=customer.id,
        status=OrderStatus.PENDING,
        total_amount=software.unit_price,
    )
    db_session.add(order)
    db_session.flush()

    line = OrderLine(
        order_id=order.id,
        product_id=software.id,
        quantity=1,
        unit_price=software.unit_price,
        line_total=software.unit_price,
        line_type=LineType.ONE_TIME,
    )
    db_session.add(line)
    db_session.commit()

    from app.services.fulfillment_service import fulfillment_service
    sugg = fulfillment_service.suggest_fulfillment(db_session, order.id)
    # Digital item should NOT require warehouse allocation
    assert len(sugg["lines"]) == 0


# =========================================================================
# 19. SERVICE PRODUCT WITHOUT WAREHOUSE REQUIREMENT
# =========================================================================
def test_19_service_product_without_warehouse_requirement(client: TestClient, db_session: Session, auth_tokens):
    srv = db_session.query(Product).filter(Product.sku == "SRV-DEPLOY-01").first()
    srv.fulfillment_type = "SERVICE"
    db_session.commit()

    customer = db_session.query(Customer).first()
    order = Order(
        order_number="ORD-TEST-SERVICE-001",
        customer_id=customer.id,
        status=OrderStatus.PENDING,
        total_amount=srv.unit_price,
    )
    db_session.add(order)
    db_session.flush()

    line = OrderLine(
        order_id=order.id,
        product_id=srv.id,
        quantity=1,
        unit_price=srv.unit_price,
        line_total=srv.unit_price,
        line_type=LineType.ONE_TIME,
    )
    db_session.add(line)
    db_session.commit()

    from app.services.fulfillment_service import fulfillment_service
    sugg = fulfillment_service.suggest_fulfillment(db_session, order.id)
    # Service item should NOT require warehouse allocation
    assert len(sugg["lines"]) == 0


# =========================================================================
# 20. HYBRID ORDER (PHYSICAL + RECURRING PRODUCTS)
# =========================================================================
def test_20_hybrid_order_physical_and_recurring(client: TestClient, db_session: Session, auth_tokens):
    customer = db_session.query(Customer).filter(Customer.company_name == "Acme Corp").first()
    laptop = db_session.query(Product).filter(Product.sku == "HW-IOT-100").first()
    software = db_session.query(Product).filter(Product.sku == "SUB-SUPP-247").first()

    laptop.fulfillment_type = "PHYSICAL"
    software.fulfillment_type = "DIGITAL"
    db_session.commit()

    order = Order(
        order_number="ORD-TEST-HYBRID-GOLDEN",
        customer_id=customer.id,
        status=OrderStatus.CONFIRMED,
        total_amount=laptop.unit_price + software.unit_price,
    )
    db_session.add(order)
    db_session.flush()

    line_phys = OrderLine(
        order_id=order.id,
        product_id=laptop.id,
        quantity=1,
        unit_price=laptop.unit_price,
        line_total=laptop.unit_price,
        line_type=LineType.ONE_TIME,
    )
    line_rec = OrderLine(
        order_id=order.id,
        product_id=software.id,
        quantity=1,
        unit_price=software.unit_price,
        line_total=software.unit_price,
        line_type=LineType.RECURRING,
        subscription_enabled=True,
        billing_frequency="MONTHLY",
        duration_mode="TILL_VALIDITY",
        validity_value=3,
        validity_unit="MONTHS",
    )
    db_session.add_all([line_phys, line_rec])
    db_session.flush()

    sub = Subscription(
        customer_id=customer.id,
        order_id=order.id,
        product_id=software.id,
        name="Mission Critical Support",
        duration_mode="TILL_VALIDITY",
        validity_value=3,
        validity_unit="MONTHS",
        billing_frequency="MONTHLY",
        status=SubscriptionStatus.ACTIVE,
        start_date=datetime.now(timezone.utc),
        end_date=datetime.now(timezone.utc) + relativedelta(months=3),
        next_billing_date=datetime.now(timezone.utc) + relativedelta(months=1),
    )
    db_session.add(sub)
    db_session.commit()

    res = client.post(
        f"/api/orders/{order.id}/billing",
        headers={"Authorization": f"Bearer {auth_tokens['finance']}"},
    )
    assert res.status_code == 200
    invoices = res.json()["invoices"]
    types = [i["billing_type"] for i in invoices]
    assert "ONE_TIME" in types
    assert "RECURRING" in types
